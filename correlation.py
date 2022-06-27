import pandas as pd
from openpyxl import load_workbook
import scipy.stats as scpy

# OSCAR: I think, overall, it is easier to process data with data frames:
# read data from the excel files into DFs, then processing it to get results as more DFs, and then writing this DFs
# to either an excel or CSV file
# RESOURCES:
# https://www.geeksforgeeks.org/python-pandas-dataframe/
# https://pandas.pydata.org/docs/user_guide/indexing.html

#
#   Retrieve Data From Analysis Tool Output
#

# Reads in warning per snippet data from each analysis tool.
# There is data for one analysis tool per excel file.
# The data frames for each file are returned in a list.
def readData():
    dfList = []

    dfList.append(pd.read_excel('data/checker_framework_data.xlsx'))

    #*** Add more analysis tool output here***

    return dfList

# Read all the data output data frames from the various analysis tool 
# and create a list of all the unique snippets across all the datasets that contain warnings
def getSnippetsWithWarnings(dfList):
    uniqueSnippets = []

    for df in dfList:
        listSnippets = df['Snippet'].to_list()
        uniqueSnippets.extend(list(set(listSnippets)))

    # Name of snippets in 'uniqueSnippets' format example: COG Dataset 1 - 12
    #                                              format: Dataset Name - Snippet #
    return uniqueSnippets

# Gets a count of the number of snippets that contain warnings for each dataset
# Returns a dictionary where the keys is the names of data sets. The values are an integer count of 
# the number of snippets that contain warnings for that data set.
def sortUniqueSnippetsByDataset(datasets, uniqueSnippets):
    countSnippetsPerDataset = dict([(key.split("-")[1].strip(), 0) for key in datasets])
 
    for snippet in uniqueSnippets:
        #TODO: Could maybe simplify all these key.splits by doing it once when the list is first created if the other data is not needed
        snippet = snippet.split("-")[0].strip() # Name of snippets in 'uniqueSnippets' format example: COG Dataset 1 - 12
                                                #                                              format: Dataset Name - Snippet #
        if snippet in countSnippetsPerDataset:
            countSnippetsPerDataset[snippet] += 1

    return countSnippetsPerDataset

# Determines the number of warnings for each snippet, separated by the dataset the snippet is from.
# Creates a dictionary where the keys are the names of data sets. Values are a list where the size is 
# the TOTAL number of snippets in the dataset and values within the list are the number of warnings for a given snippet.
def getNumWarningsPerSnippet(dfList, numSnippetsJudgedPerDataset, datasets):
    warningsPerSnippetPerDataset = dict([(key.split('-')[1].strip(), 0) for key in datasets])

    # Setup the dictionary with empty lists
    count = 0
    for dataset in warningsPerSnippetPerDataset:
        warningsPerSnippetPerDataset[dataset] = [0] * int(numSnippetsJudgedPerDataset[count])
        count += 1

    # Loop through the analysis tool output dataframes
    for df in dfList:
        numWarnings = df.sum(axis=1, numeric_only=True).tolist()
        snippetNames = df['Snippet'].to_list()

        print(numWarnings)

        if len(snippetNames) != len(numWarnings):
            raise Exception("Number of snippets does not match number of warnings associated with said snippets") 

        for i in range(len(snippetNames)):
            snippetDataset = snippetNames[i].split('-')[0].strip()
            snippetNumber = snippetNames[i].split('-')[1].strip()

            #********************Temporary if statement
            #TODO change name of the fMRI snippets to match this standard i.e. "fMRI Dataset - 1"
            if snippetNumber.isnumeric():
                warningsPerSnippetPerDataset[snippetDataset][int(snippetNumber) - 1] += numWarnings[i]

    print(warningsPerSnippetPerDataset)

    return warningsPerSnippetPerDataset

#
#   Retrieve Data From Studies
#

# Reads the results of the cog data set 3 study. It contains 120 people who rated 100 snippets on a scale of 1-5.
# 1 being less readable and 5 being more readable.
# OSCAR: where are we filtering out the 4 snippets that are commented out?
# OSCAR: in cog_dataset_3.csv, are the snippets identified by column index?
def getAveragesCogDataset3():
    df = pd.read_csv('data/cog_dataset_3.csv')

    # Returns a list of the averages for each snippet
    return [round(sum(df[column]) / len(df[column]), 2) for column in df.columns[2:]]

#
#   Interact with datapoints.xlsx
#

# Takes the averages of the complexity metric and places them into the 
# first column of dataapoints.xlsx sheet cog_dataset_3.
def setCogDataset3ComplexityMetricColumn(complexityMetrics):
    workbook = load_workbook('data/datapoints.xlsx')
    worksheet = workbook['cog_dataset_3']

    for i, value in enumerate(complexityMetrics):
        worksheet.cell(row=i + 1, column=1, value=value)

    # possible dataframe syntax (not sure)
    # df = read_excel_file(...)
    # df[:5]
    # x = df["Dataset"]

    workbook.save('data/datapoints.xlsx')

# Takes a dictionary of the following format: Keys are the names of the datasets. Values are a list where the size is 
# the TOTAL number of snippets in the dataset and values within the list are the number of warnings for a given snippet.
# Indexing the list references a snippet by its number - 1.
def setCogDataset3WarningCountColumn(warningsPerSnippetPerDataset):
    workbook = load_workbook('data/datapoints.xlsx')

    #COG DATASET 3

    worksheet = workbook['cog_dataset_3']

    for i, value in enumerate(warningsPerSnippetPerDataset['COG Dataset 3']):
        worksheet.cell(row=i + 1, column=2, value=value)



    workbook.save('data/datapoints.xlsx')

#
#   Interact with correlation_analysis.xlsx
#

# Gets a list of all the datasets that snippets can come from
def getDatasets():
    df = pd.read_excel('data/correlation_analysis.xlsx', usecols=['Complexity Metric'])

    return df['Complexity Metric'].to_list()

def getNumSnippetsJudgedColumn():
    df = pd.read_excel('data/correlation_analysis.xlsx', usecols=['# of snippets judged (complexity)'])

    return df['# of snippets judged (complexity)'].to_list()

# Sets the values of the column "# of snippets with warnings" in the correlation analysis excel sheet
def setNumSnippetsWithWarningsColumn(countSnippetsPerDataset):
    workbook = load_workbook('data/correlation_analysis.xlsx')
    worksheet = workbook.active # gets first sheet

    for i, value in enumerate(countSnippetsPerDataset.values()):
        # Writes a new value PRESERVING cell styles.
        worksheet.cell(row=i + 2, column=3, value=value)

    workbook.save('data/correlation_analysis.xlsx')

def setNumDatapointsForCorrelationColumn():
    workbookDatapoints = load_workbook('data/datapoints.xlsx')

    workbookAnalysis = load_workbook('data/correlation_analysis.xlsx')
    worksheetAnalysis = workbookAnalysis.active
    
    # Loop through every sheet in datapoints.xlsx. A sheet corresponds to the datapoints for a specific dataset
    for i, worksheetDatapoints in enumerate(workbookDatapoints.worksheets):
        numDataPoints = worksheetDatapoints.max_row

        worksheetAnalysis.cell(row=i + 2, column=4, value=numDataPoints)

    workbookAnalysis.save('data/correlation_analysis.xlsx')

def setKendallTauColumn(kendallTauVals):
    workbook = load_workbook('data/correlation_analysis.xlsx')
    worksheet = workbook.active

    for i, value in enumerate(kendallTauVals):
        worksheet.cell(row=i + 2, column=5, value=value)

    workbook.save('data/correlation_analysis.xlsx')

#
#   CORRELATION
#
# OSCAR: probably with dataframes and slices, the code would be simpler
# OSCAR: we need to also compute Spearman Correlation, which doesn't assume normal distributions (scipy.stats.spearmanr)
def kendallTau():
    workbookDatapoints = load_workbook('data/datapoints.xlsx')

    kendallTauVals = []

    # Loop through every sheet in datapoints.xlsx. A sheet corresponds to the datapoints for a specific dataset
    for i, worksheetDatapoints in enumerate(workbookDatapoints.worksheets):

        for row in worksheetDatapoints.iter_rows():
            x.append(row[0].value)
            y.append(row[1].value)

        corr, pValue = scpy.kendalltau(x, y)

        kendallTauVals.append(corr)

    return kendallTauVals

if __name__ == '__main__':
    dfList = readData()
    datasets = getDatasets()

    setNumSnippetsWithWarningsColumn(sortUniqueSnippetsByDataset(datasets, getSnippetsWithWarnings(dfList)))

    # Already set, no need to run multiple times
    #setCogDataset3ComplexityMetricColumn(getAveragesCogDataset3())

    #getNumWarningsPerSnippet(dfList, getNumSnippetsJudgedColumn(), datasets)
    setCogDataset3WarningCountColumn(getNumWarningsPerSnippet(dfList, getNumSnippetsJudgedColumn(), datasets))

    setNumDatapointsForCorrelationColumn()

    setKendallTauColumn(kendallTau())